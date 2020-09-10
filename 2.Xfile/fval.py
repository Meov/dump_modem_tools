#!/usr/bin/env
import sys
import os
import binascii
from openpyxl import Workbook
from openpyxl import load_workbook

class ConVerter:
    __tmp     = []
    __rows    = 0
    __columns = 0
    symToaddr = {}
    
    def __init__(self,symfile):
        self.symfile = symfile

        # a new book
        book = load_workbook(self.symfile)
        print(book)
        self.book = book
        # we just need the sheet1 
        sheet1 = book['Sheet1']
        self.sheet = sheet1
        print(self.sheet)
        print(self.sheet.max_row)
        __rows = self.sheet.max_row
        print(self.sheet.max_column)
        __columns = self.sheet.max_column

    def __listpickup__(self,nm):
        #__tmp = list(self.sheet.columns)

        if (nm > self.sheet.max_column):
            return -1

        lst = []

        for r in range(1,self.sheet.max_row):
            v = self.sheet.cell(r,nm).value
            lst.append(v)

        return lst

    def databuild(self,ids=[]):
        __tmp = list(self.sheet.columns)
        print(ids)
        targetlst = []
       
        keylist = self.__listpickup__(ids[0])

        for i in range(1,len(ids)):
            if(ids[i] > self.sheet.max_column):
                print('error')
                return -1

            targetlst.append(self.__listpickup__(ids[i]))

        a = zip(*targetlst)
        self.info = dict(zip(keylist,a))
        return self.info

