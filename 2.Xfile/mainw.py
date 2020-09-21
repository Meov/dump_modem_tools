#!/usr/bin/env

import sys
import os
import subprocess
import struct
import re
import codecs
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from fval import ConVerter
from mem import SymToMem
from PyQt5.QtWidgets import QMainWindow,QDialog,QWhatsThis
from PyQt5.QtWidgets import QFileDialog,QMessageBox
from PyQt5.QtCore import QTimer,QThread,pyqtSignal,QEvent
from PyQt5.QtGui import QIcon
from Ui_mainwindow import *       #main window
from Ui_querydialog import *    #query dialog
from Ui_memory import *         #memory window
from bin_creat import BinCreat

VERSION = "VERSION 0.0.5"
name_bin_format = ".bin"
forsym = ['g_exc_desc_arr','s_exc_stat']
regsym = ['g_exc_svc_arr', \
        'g_exc_sys_arr', \
        'g_exc_reserve_arr', \
        'g_exc_abort_arr', \
        'g_exc_undef_arr',\
        'g_exc_fiq_arr']

RstReason = ['Data Abort', \
            'Address 0x0 Jump Exception', \
            "Prefech Abort()", \
            'Undefined Instuction Abort ', \
            'TP_OS_ASSERT(0),assert failed ', \
            'Thread stack overflow ', \
            'unknown reason found !',\
            ]
work_mode = [
    'USER MODE',\
    'FIQ MODE',\
    'IRQ MODE',\
    'Supervisor MODE',\
    'Abort MODE',\
    'Undefined MODE',\
    'System MODE'
]

bincreat = BinCreat()
stack_flag = [0,0,0,0,0,0,0,4294967295]

str_not_init = 'not_initialized'
name_tar_format = ".tar.bz2"
class DumpGetThread(QtCore.QThread):
    #  define signal
    _signal = pyqtSignal(str)
 
    def __init__(self,dump_file):
        super(DumpGetThread, self).__init__()
        self.dump_file = dump_file
        
    def __del__(self):
        print("exit")

    def run(self):
        if  self.dump_file == '':
            return 
        dump_split_file_dir = bincreat.get_file_dir(self.dump_file[1])
        dump_bin_save_dir = bincreat.get_file_dir(self.dump_file[2])


        if(name_bin_format in self.dump_file[0]):  #.bin文件
            bin_file_path = os.path.join(dump_split_file_dir,self.dump_file[0])
            dump_file_path_str = bin_file_path.replace('\\','/')   
            self._signal.emit(dump_file_path_str)
            return

        if (os.path.isdir(dump_split_file_dir) and (os.path.isdir(dump_bin_save_dir))):
            if(self.dump_file[0]):
                #print(dump_split_file_dir)
                print(self.dump_file[0])
                print(dump_bin_save_dir)

                bin_file_path = os.path.join(dump_split_file_dir,self.dump_file[0]+".bin")
                if os.path.exists(bin_file_path):   #delete existed files
                    os.remove(bin_file_path)
                bincreat.binary_file_create(dump_split_file_dir,dump_bin_save_dir, self.dump_file[0])

                #bin_file_path = os.path.join(dump_split_file_dir,self.dump_file[0]+".bin")
                dump_file_path_str = bin_file_path.replace('\\','/')              
                #self._signal.emit("created_ok")  # 注意这里与_signal = pyqtSignal(str)中的类型相同
                
                self._signal.emit(dump_file_path_str)


            else:
                self.printXfile("could not find dumped files!")
                self.statusbar.showMessage("err: could not find dumped files!")
                return
        else:
            self.printXfile("path not right!")
            return

class MemoryShowDialog(QDialog,Ui_Dialog_Mem):
    def __init__(self,parent=None):
        super(MemoryShowDialog,self).__init__(parent)
        self.setupUi(self)
        self.setWindowIcon(QIcon("./X.ico"))
        self.setFixedSize(self.width(),self.height())
        self.memory_content.setText(" ")
        self.mem_content_cancel.clicked.connect(self.cancle_dialog)
        self.mem_content_clear.clicked.connect(self.clear_dialog)

    def printMfile(self,toText):
        self.memory_content.append(toText)
        self.cursor=self.memory_content.textCursor()
        self.memory_content.moveCursor(self.cursor.End)
        # was suggested for smooth showing
        QtWidgets.QApplication.processEvents()
    def clear_dialog(self):
        self.memory_content.clear()
    def cancle_dialog(self):
        self.close()

class QueryDialog(QDialog,Ui_Dialog_Query):
    expression_signal = QtCore.pyqtSignal(str)
    addr_signal = QtCore.pyqtSignal(str)

    def __init__(self,parent=None):
        super(QueryDialog,self).__init__(parent)
        self.setupUi(self)
        self.setWindowIcon(QIcon("./X.ico"))
        self.setFixedSize(self.width(),self.height())
        self.Query_Button.clicked.connect(self.lookup_line_value)
        self.Query_addr_Button.clicked.connect(self.lookup_addr_line)
    def lookup_line_value(self):
        self.expression_signal.emit(str(self.ValueQuery_line.text()))

    def lookup_addr_line(self):
        addr =  self.ValueQureyAddr_line.text()
        addr_size = self.QueryAddrLen_line.text()
        #memory size is set to a multiple of 4 bytes as U32
        if not addr_size:
            addr_size = 512 #defualt 512 bytes
        else:
            try:
                addr_size = int(addr_size)
            except:
                QMessageBox.warning(self,"error","integer only!",QMessageBox.Ok) 
                return
            if(addr_size%4 != 0):   #Round up by 4
                addr_size = addr_size//4*4 + 4
    
        addr_info = addr+":"+str(addr_size)
        print(addr_info)
        self.addr_signal.emit(str(addr_info))

    def event(self, event):
        if event.type()==QEvent.EnterWhatsThisMode:
            QWhatsThis.leaveWhatsThisMode()
            QMessageBox.warning(self,"版本",VERSION,QMessageBox.Ok)
        return QDialog.event(self,event)

class MainWindow(QMainWindow,Ui_MainWindow):

#initialize
    def __init__(self,parent=None):
        super(MainWindow,self).__init__(parent)
        self.setupUi(self)
        #self.Title.setText("X-file generator")
        font = QtGui.QFont()
        
        self.setWindowIcon(QIcon("./X.ico"))
        
        font.setFamily("Microsoft YaHei")
        self.xfileoutput.setFont(font)
        self.dumpfile = ''
        self.symfile = ''
        self.stm = str_not_init
        self.asfPath = 'L1860-MODEM.axf'
        self.add2linePath = './addr2line.exe'
        self.readelfPath = './readelf.exe'
        self.symButton.clicked.connect(self.openSymfile)
        self.generateButton.clicked.connect(self.generateXfile)
        self.txtButton.clicked.connect(self.saveTotxt)
        self.actionSelectFile.triggered.connect(self.selectAsf)
        self.actionToSymsTbl.triggered.connect(self.toSymbolTables)
        self.actionSelectExe.triggered.connect(self.selectAddr2line)
        self.generateButton.setEnabled(False)
        
        #registers
        self.sysval = []
        self.usrval = []
        self.sysval = []
        self.fiqval = []
        self.abortval=[]
        self.undefval=[]


        #dialog QueryDialog and MemoryShowDialog
        self.actionquery_item.triggered.connect(self.query_dialog_action)        
        self.query_dialog = QueryDialog()
        self.dump_browse_Button.clicked.connect(self.open_bin_file)
        self.memory_window = MemoryShowDialog()
        self.query_dialog.expression_signal.connect(self.get_expression_data)  #_siganl set
        self.query_dialog.addr_signal.connect(self.get_addr_data)

        #current directory default
        self.dump_split_file_dir = '.'
        self.dump_bin_save_dir = '.'

        #progressbar
        self.progressBar.hide()
        self.timer = QTimer()
        self.timer.timeout.connect(self.update_progress)
        self.progress_value = 0
        self.progressBar.setValue(self.progress_value)

#query memory bin data
    def query_dialog_action(self):
        self.query_dialog.show()
    def memory_window_printmem(self,data,data_info):
        self.memory_window.show()
        self.memory_window.printMfile(" ")
        self.memory_window.printMfile(data_info)
        self.memory_window.printMfile(" ")
        for i in range(len(data)//4):
            self.memory_window.printMfile('0x'+'{0:0>8x} '.format(data[0+i*4])+ \
            '0x'+'{0:0>8x} '.format(data[1+i*4])+ \
            '0x'+'{0:0>8x} '.format(data[2+i*4])+ \
            '0x'+'{0:0>8x} '.format(data[3+i*4]))
        
        for i in range(len(data)%4):
            self.memory_window.printMfile('0x'+'{0:0>8x} ' .format(data[len(data)//4 *4 + i]))

    def get_addr_data(self,parameter):  
        addr_info = parameter.split(':')
        addr = addr_info[0]
        size = int(addr_info[1])
        
        if(addr ==' '):
            print("no input data!")
            QMessageBox.warning(self,"error","no input data!",QMessageBox.Ok) 
            return
        if(self.stm == str_not_init):
            QMessageBox.warning(self,"error","mem data NOT initialized!!",QMessageBox.Ok) 
            return
        
        try:    #check addr is hex format
            addr = int(addr,16)
        except:
            print("enter HEX only!")
            QMessageBox.warning(self,"error","HEX only!",QMessageBox.Ok)
            return
        try:    #try to get memory data based on base addr and size    
            val = self.stm.toMemContent(addr,size)
            if val == '':
                self.memory_window.printMfile("can't find: "+'0x'+'{0:0>8x} '.format(addr))
                print("can't find :"+'0x'+'{0:0>8x} '.format(addr))
                return
            datatype = "U32"
            data = []
            if datatype == "U32":
                for i in range(size//4):
                    tmp = int.from_bytes(val[i*4:(i*4+4)],'little')
                    data.append(tmp)
            print(data)
            data_info = "----Addr: "+ '0x'+'{0:0>8x} '.format(addr)+"    size:"+str(size)+"----"
            self.memory_window_printmem(data,data_info)
        except:
            print("can't find :"+'0x'+'{0:0>8x} '.format(addr))
            QMessageBox.warning(self,"error","can't find: "+'0x'+'{0:0>8x} '.format(addr),QMessageBox.Ok)   
           
    def get_expression_data(self,parameter):
        if(parameter == ''):
            print("no input data!")
            QMessageBox.warning(self,"error","no input data!",QMessageBox.Ok) 
            return
        if(self.stm == str_not_init):
            print("mem data not initialized!")
            QMessageBox.warning(self,"error","mem data NOT initialized!!",QMessageBox.Ok) 
            return
        try:
            data = self.get_symbol_meminfo(self.stm,parameter,"unknown")
            
            data_info = "---Expression: "+parameter+"    Size: " \
                +str(data[1])+"    Addr: "+ '0x'+'{0:0>8x} '.format(data[0])+"----"
            self.memory_window_printmem(data[2],data_info)
            #print("the value of "+ parameter + " is: ")
            #print(data)
        except:
            print("can't find: "+parameter)
            QMessageBox.warning(self,"warning","can't find: "+parameter,QMessageBox.Ok) 

#Xfile creation
    def selectAsf(self):
        asfFile = QFileDialog.getOpenFileName(None, 'Open file', '.','axf file(*.axf *.elf)')
        if asfFile:
            self.asfPath = asfFile[0]
            self.statusbar.showMessage(self.asfPath)
            self.printXfile("loaded: "+self.asfPath)

    def symtxtPreProcess(self,filename):
        if os.path.exists(filename):
            f_old = open(filename,mode='r+')
            f_new = open('new_'+filename,'w+')
            next(f_old)
            next(f_old)

            for line in f_old:
                line = re.sub(' +','\t',line)
                f_new.write(line)
            f_old.close
            f_new.close
            
        else:
            self.printXfile("symbol file not found ~~~")

    def toSymbolTables(self):
        readelfFile = QFileDialog.getOpenFileName(None, 'Open file', '.','fromelf exe file(*.exe)')
        if readelfFile:
            
            self.readelfPath = readelfFile[0]
            self.statusbar.showMessage(self.readelfPath)
            tblfilepath = "symbols.txt"
            execute = [self.readelfPath,'-s',self.asfPath]
            try:
                run = subprocess.Popen(execute,stdin=subprocess.PIPE,stdout=subprocess.PIPE, \
                    stderr=subprocess.STDOUT,bufsize=0)
                out = run.communicate()
                f = open(tblfilepath,'wb+')
                f.write(out[0])
                f.close
                self.symtxtPreProcess(tblfilepath)
                self.txt_to_xlsx('new_'+tblfilepath,'sym.xlsx')
            except:
                return 
            
    def selectAddr2line(self):
        add2lineFile = QFileDialog.getOpenFileName(None, 'Open file', '.','addr2line(*.exe)')
        self.add2linePath = add2lineFile[0]
        self.statusbar.showMessage(self.add2linePath)
        print("loaded: "+self.add2linePath)
        self.printXfile("loaded: "+self.add2linePath)
             
    def txt_to_xlsx(self,filename,outfile):
        if(os.path.exists(filename)):
            fr = codecs.open(filename,'r')
            wb = openpyxl.Workbook()
            ws = wb.active
            ws = wb.create_sheet()
            ws.title = 'Sheet1'
            row = 0
            for line in fr:
                row +=1
                line = line.strip()
                line = line.split('\t')
                col = 0
                for j in range(len(line)):
                    col +=1
                    #print (line[j])
                    ws.cell(column = col,row = row,value = line[j].format(get_column_letter(col)))
            wb.save(outfile)
            self.printXfile("new symbols file generated!")
        else:
            self.printXfile("symbol file not found ~~~")

    def initialize(self,symfile,dumpfile):
        filename = symfile
        memfile  = dumpfile
        ids = [8,2,3]
        cvt = SymToMem(filename,memfile,ids)
        return cvt    

    def get_symbol_meminfo(self,SymToMem,symbol,datatype=''):
        address = SymToMem.toMemaddress(symbol)
        size    = SymToMem.toMemSize(symbol)
        data    = SymToMem.toMemContent(address,size)
        val = []

        if datatype == "U32":
            for i in range(size//4):
                #val += int(struct.unpack('I',data[i*4:(i*4+4)]))
                tmp = int.from_bytes(data[i*4:(i*4+4)],'little')
                #tmp = int(tmp,base=16)
                val.append(tmp)
        
        #uncertain the size of varibles queried in the dialog
        elif(datatype == "unknown"):
            u32_num = 0
            for i in range(size//4): 
                tmp = int.from_bytes(data[i*4:(i*4+4)],'little')
                val.append(tmp)
                u32_num = i
            if(size%4 != 0):
                tmp = int.from_bytes(data[u32_num*4:u32_num*4+size%4],"little")
                val.append(tmp)

        result = [address,size,val]
        print(result)
        return result

    def print_mem(self,data,size,datatype=''):
        val = []
        if datatype == 'U32':
            for i in range(size//4):
                #print(i)
                val += struct.unpack('I',data[i*4:(i*4+4)])
        for i in range(size//4):
            print("0x%x " %val[i]),

    def get_reset_reason(self,SymToMem):

        #symToaddr = cvt.databuild(ids)

        #address = SymToMem.toMemaddress(forsym[1])
        #size    = SymToMem.toMemSize(forsym[1])
        #data = SymToMem.toMemContent(address,size)

        #data = struct.unpack('I',data[:size])
        data = self.get_symbol_meminfo(SymToMem,forsym[1],'U32')

        rstval = data[2][0]
        print(rstval)
        if rstval != 205:
            reason = "not a reset"
            f = open('xfile.txt','w+')
            self.printXfile('{:0>X}'.format(rstval)+"...not a reset...")
            f.close()
            print('No Reset found ...')
            #exit(0) 

        #Step 2: find reset reason
        #address = SymToMem.toMemaddress(forsym[0])
        #size    = SymToMem.toMemSize(forsym[0])
        #data = SymToMem.toMemContent(address,size)
        #data = struct.unpack('III',data[0:size])
        ret = self.get_symbol_meminfo(SymToMem,forsym[0],'U32')   #判断死机类型
        print(ret)
        data = ret[2]
        if data[2] == 4:
            reason = RstReason[0]
        elif data[2] == 0:
            reason = RstReason[1]
        elif data[2] == 1:
            reason = RstReason[2]
        elif data[2] == 2:
            reason = RstReason[3]
        elif data[2] == 131:
            reason = RstReason[4]   #assert
        elif data[2] == 224:
            reason = RstReason[5]
        else:
            reason = "unknown reason found !"
            
        print("reset No. is %x : " % data[2]),
        print(reason)
        return reason

    def get_work_mode(self,reg_cpsr):
        mode = reg_cpsr&0x1f
        mode_str = ""
        if(mode == 0x10):
            print("USER")
            mode_str = work_mode[0]
        elif(mode == 0x11):
            print("FIQ")
            mode_str = work_mode[1]
        elif(mode == 0x12):
            print("IRQ")
            mode_str = work_mode[2]
        elif(mode == 0x13):
            print("Supervisor")
            mode_str = work_mode[3]
        elif(mode == 0x17):
            print("Abort")
            mode_str = work_mode[4]
        elif(mode == 0x1B):
            print("Undefined")
            mode_str = work_mode[5]
        elif(mode == 0x1F):
            print("System")
            mode_str = work_mode[6]
        return mode_str

    def get_register_from_mode(self,current_work_mode):
        if current_work_mode == work_mode[1]: #fiq
            r13 = self.fiqval[2][5]
            r14 = self.fiqval[2][6]
            r15 = self.fiqval[2][6]
            print(current_work_mode)
        elif current_work_mode == work_mode[6]: #sys
            r13 = self.sysval[2][0]
            r14 = self.sysval[2][1]
            r15 = self.sysval[2][1]
            print(current_work_mode)
        elif current_work_mode == work_mode[4]: #abt
            r13 = self.abortval[2][0]
            r14 = self.abortval[2][1]
            r15 = self.abortval[2][1]
            print(current_work_mode)
        elif current_work_mode == work_mode[5]:  #udf
            r13 = self.undefval[2][0]
            r14 = self.undefval[2][1]
            r15 = self.undefval[2][1]
            print(current_work_mode)
        elif current_work_mode == work_mode[2]: #irq
            r13 = self.undefval[2][0]
            r14 = self.undefval[2][1]
            r15 = self.undefval[2][1]
            print(current_work_mode)
        else:                                   #svc
            r13 = self.usrval[2][0]
            r14 = self.usrval[2][1]
            r15 = self.usrval[2][2]
            print(current_work_mode)

        return [r13,r14,r15]

    def get_cpu_registers(self,SymToMem,rst_reason):
        #g_exc_svc_arr
        self.usrval  = self.get_symbol_meminfo(SymToMem,regsym[0],'U32')
        #g_exc_reserve_arr
        self.usrresv = self.get_symbol_meminfo(SymToMem,regsym[2],'U32')
        #g_exc_sys_arr
        self.sysval = self.get_symbol_meminfo(SymToMem,regsym[1],"U32")
        #g_exc_fiq_arr
        self.fiqval = self.get_symbol_meminfo(SymToMem,regsym[5],"U32")   
        #g_exc_abort_arr
        self.abortval = self.get_symbol_meminfo(SymToMem,regsym[3],"U32")
        #g_exc_undef_arr
        self.undefval = self.get_symbol_meminfo(SymToMem,regsym[4],"U32")

        regs=[]
        #fill r[0]-r[12]
        for i in range(0,13):
            regs.append(self.usrresv[2][i])

        #R13 R14 R15
        #R15(PC) = R14(LR)
        if rst_reason == RstReason[0]:  #Data Abort
            #g_exc_abort_arr
            work_mode_str = self.get_work_mode(self.abortval[2][2])
            r13_15 = self.get_register_from_mode(work_mode_str)
            regs.append(r13_15[0])
            regs.append(r13_15[1])
            regs.append(self.abortval[2][1]-8)   #r15 ==>PC
        if rst_reason == RstReason[2]:    #"Prefech Abort ()"
            #g_exc_abort_arr
            work_mode_str = self.get_work_mode(self.abortval[2][2])
            r13_15 = self.get_register_from_mode(work_mode_str)
            regs.append(r13_15[0])
            regs.append(r13_15[1])
            regs.append(self.usrval[2][1]-4)     #abortval_r14 is broken
        if rst_reason == RstReason[1]:  # Address 0x0 Jump Exception 
            work_mode_str = self.get_work_mode(self.usrresv[2][13])
            r13_15 = self.get_register_from_mode(work_mode_str)
            regs.append(r13_15[0])
            regs.append(r13_15[1])
            regs.append(r13_15[2])
        if rst_reason == RstReason[3]:  #undefine err
            work_mode_str = self.get_work_mode(self.undefval[2][2])
            r13_15 = self.get_register_from_mode(work_mode_str)
            regs.append(r13_15[0])
            regs.append(r13_15[1])      
            regs.append(self.undefval[2][1])-4

        if rst_reason == RstReason[4]:  #assert
            work_mode_str = self.get_work_mode(self.usrresv[2][13])
            r13_15 = self.get_register_from_mode(work_mode_str)
            regs.append(r13_15[0])
            regs.append(r13_15[1])
            regs.append(r13_15[2])      
        if rst_reason == RstReason[5]:  # Thread stack overflow
            work_mode_str = self.get_work_mode(self.usrresv[2][13])
            print("Thread stack overflow")
        for i in range(len(regs)):
            print("Reg[%d] = 0x%x" %(i,regs[i]))
        rst_reason = rst_reason+'-----'+work_mode_str+'------'
        return (regs,rst_reason)

    def get_stack_content(self,SymToMem,spaddr):
        block =[0]
        size=0
        
        while block[size] != 4294967295 and size < 1000:
            #print(spaddr)
            data = SymToMem.toMemContent(spaddr,4)
            spaddr = spaddr+4
            #block += struct.unpack('I',data[:4])
            block.append(int.from_bytes(data,'little'))
            size = size+1

        return block[1:]

    def get_callstack(self,stack,pc):
        func_addr = [pc]
        i = 0
        retline=[]
        if(os.path.isfile(self.add2linePath) == False):
            self.printXfile('add2line not working')
            return("add2line not working")
        for i in range(len(stack)):
            if (stack[i]>72351744) and (stack[i]<73400320):
                func_addr.append(stack[i])
        print(func_addr)

        for i in range(len(func_addr)):
            #execute = ['addr2line.exe','-e','L1860-MODEM.axf','-f','-s',hex(func_addr[i])]
            execute = [self.add2linePath,'-e',self.asfPath,'-f','-s',hex(func_addr[i])]
            run = subprocess.Popen(execute,stdin=subprocess.PIPE,stdout=subprocess.PIPE, \
                stderr=subprocess.STDOUT,bufsize=0)
            out = run.communicate()
            retline.append(out)
            #print(out)
            #for line in iter(run.stdout.readline, b''):
            #    print line,
        return retline

    def openSymfile(self):
        _translate = QtCore.QCoreApplication.translate
        self.symfile = QFileDialog.getOpenFileName(None, 'Open file', '.','symbol files(*.xlsx)')
        self.symline.setText(_translate("MainWindow", self.symfile[0]))
        print(self.symfile)

    def printXfile(self,toText):
        self.xfileoutput.append(toText)
        self.cursor=self.xfileoutput.textCursor()
        self.xfileoutput.moveCursor(self.cursor.End)
        # was suggested for smooth showing
        QtWidgets.QApplication.processEvents()
    
    def generateXfile(self):
        if self.dumpfile != "tar_file":
            self.analyseXdata(self.dumpfile)
        else:   #need extract .tar.bz2 files
            self.progressBar.setValue(self.progress_value)
            self.dump_split_file_dir = bincreat.get_file_dir(self.dump_split_file_dir)
            self.dump_bin_save_dir = bincreat.get_file_dir(self.dump_split_file_dir)
            
            if (os.path.isdir(self.dump_split_file_dir) and (os.path.isdir(self.dump_bin_save_dir))):
                file_name = self.combo_dump_name.currentText()
                dump_file = [file_name,self.dump_split_file_dir,self.dump_bin_save_dir]
                self.statusbar.showMessage('generating a xfile ...')
                self.printXfile('extractig bin file '+file_name+" ...")
                self.timer.start(50)  #50ms timeout to update progress bar
                self.thread = DumpGetThread(dump_file)
                self.thread._signal.connect(self.callback_getdump)
                self.thread.start()
                self.progressBar.show()
            else:
                self.printXfile("path not right!")
                return

    def analyseXdata(self,dump_bin_path):
        self.statusbar.showMessage('generating a xfile ...')
        
        if ((self.symfile == '') or ((".xlsx" in self.symfile[0])==False)):
            self.printXfile('no input symbol files...')
            return
        if(os.path.isfile(self.asfPath) == False):
            self.printXfile('no axf file ...')
            return 
        if(os.path.isfile(self.add2linePath) == False):
            self.printXfile('add2line not working...')
            return
        if (dump_bin_path == ''):
            self.printXfile('no input  dump files...')
            return

        self.stm = self.initialize(self.symfile[0],dump_bin_path)
        print(self.symfile[0])
        print(dump_bin_path)      
        self.xfileoutput.clear()     
        rst_reason = self.get_reset_reason(self.stm)
    
        ret = self.get_cpu_registers(self.stm,rst_reason)
        if not ret:
            return
        
        Reg = ret[0]
        rst_reason = ret[1]

        self.printXfile('-----------Reset Type-----------\n'+'\n')
        self.printXfile('Reason : '+rst_reason+'\n'+'\n'+'\n')
        self.printXfile('-----------Current Registers-----\n'+'\n')
    
        for i in range(len(Reg)):
            self.printXfile("Reg["+'{:>2d}'.format(i)+"] = "+'0x'+'{:0>8x}'.format(Reg[i])+'\n' )

        self.printXfile('-----------Memory of Stack------------\n'+'\n')
        cnt = self.get_stack_content(self.stm,Reg[13])
        self.printXfile('SP point at 0x%x : \n\n' %Reg[13])
        for i in range(len(cnt)//4):
            self.printXfile('0x'+'{0:0>8x} '.format(cnt[0+i*4])+ \
                '0x'+'{0:0>8x} '.format(cnt[1+i*4])+ \
                '0x'+'{0:0>8x} '.format(cnt[2+i*4])+ \
                '0x'+'{0:0>8x} '.format(cnt[3+i*4]))

        for i in range(len(cnt) % 4):
            self.printXfile('0x'+'{0:0>8x} '.format(cnt[len(cnt)//4*4+i]))
        self.printXfile('\n')

        self.printXfile('-----------Possible Callstack-----------\n'+'\n')
        run = self.get_callstack(cnt,Reg[15])
        print(run)
        if run == "add2line not working":
            self.printXfile(run)
            return

        for i in range(len(run)):
            self.printXfile("%d : " %i),
            self.printXfile(run[i][0].decode().replace('\r',''))
        self.printXfile('\n---------------End of Xfile---------------\n')

    def saveTotxt(self):
        f = open('xfile.txt','w+')
        strTxt = self.xfileoutput.toPlainText()
        fmtTxt = str(strTxt)
        print(f.write('{}'.format(fmtTxt)))
        f.close()   

#get data
    def open_bin_file(self):
        
        self.combo_dump_name.clear()
        self.dumpfile = " "
        dump_files = QFileDialog.getOpenFileNames(self,"open file",'.','dump files(*.tar.bz2 *.bin)')
        if not dump_files[0]:
            print("No bin file")
            self.printXfile( "No .bin or .tar.bz2 files!")
            return
        
        self.dump_split_file_dir, file_name = os.path.split(dump_files[0][0])
        self.dump_path_line.setText(str(self.dump_split_file_dir))
        
        file_name = file_name.split('.',2)
        
        if((name_bin_format in dump_files[0][0]) and (len(dump_files[0]) == 1)):    #found bin file  
            self.combo_dump_name.addItem(file_name[0])
            self.dumpfile = dump_files[0][0]
            self.generateButton.setEnabled(True)
        else:
            try:
                file_name = file_name[0].split('-',2)
                file_name = file_name[0]+'-'+file_name[1]
                print(file_name)
                self.dumpfile = "tar_file"
                self.combo_dump_name.addItem(file_name)
                self.generateButton.setEnabled(True)
            except:
                print("split files must named as: XXX-X-X")
                self.printXfile( "split files must named as: XXX-X-X")
                return

        

    def callback_getdump(self,parameter): 
        self.progress_value = 100
        self.progressBar.setValue(int(self.progress_value))
        self.progress_value = 0
        self.timer.stop()
        self.printXfile("Created  file ok, saved in: "+ parameter)
        print(parameter)
        self.progressBar.hide()
        self.analyseXdata(parameter)

#closeEvent 
    def closeEvent(self, event):  #overite closeEvent
        sys.exit()

#progressBar    
    def update_progress(self):
        self.progress_value += 1
        if(self.progress_value >= 99):
            self.timer.stop()
            self.progress_value = 99
        self.progressBar.setValue(int(self.progress_value))
