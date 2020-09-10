#!/usr/bin/env
import sys
import os
import binascii
from openpyxl import Workbook
from openpyxl import load_workbook

if len(sys.argv)<2:
    filename = "sym.xlsx"    
else:
    filename = sys.argv[1]
print(filename)

if len(sys.argv)==3:
    forsym = sys.argv[2]
else:
    forsym = 'g_exc_desc_arr'
print(type(forsym))

# a new book
book = load_workbook(filename)
print(book)

# we just need the sheet1 
sheet1 = book['Sheet1']

tmp = list(sheet1.columns)

print(sheet1.max_row)
rows = sheet1.max_row
print(sheet1.max_column)
columns = sheet1.max_column

addrlist = [0]

for r in range(1,rows):
    v = sheet1.cell(r,2).value
    addrlist.append(v)
#print addrlist
print(type(addrlist[2]))

symlist = [0]
for r in range(1,rows):
    s = sheet1.cell(r,8).value
    symlist.append(s)
#print symlist
print(type(symlist[4]))

sizelist = [0]
for r in range(1,rows):
    s = sheet1.cell(r,3).value
    sizelist.append(s)

temp = zip(addrlist,sizelist)
symToaddr = dict(zip(symlist,temp))

#print(symToaddr[forsym])
#print symToaddr
print(type(symToaddr[forsym][0]))
print(symToaddr[forsym][0])
print(forsym+":")
a = symToaddr[forsym][0]
print(type(a))
val = 0
if type(a) == int:
    val = int(str(a),base=16)
    print(hex(val))
if type(a) == str:
    for i in range(len(a)):
        val = val*16
        if '0' <= a[i] <= '9':
            val = val+int(a[i])
        if 'a' <= a[i] <= 'f':
            tmp = ord(a[i])-ord('a')+10
            val += tmp   
    print(type(val))    
    print(hex(val))
size = symToaddr[forsym][1]
print(size)
print(type(size))

filebin = open('dump.bin',mode='rb')
print(filebin)
filebin.seek(val,0)
print(filebin.tell())
dd = filebin.read(size)
for i in range(size):
    #print(''),
    print('%x' % dd[i]),
print(dd)

filebin.close()
#print(symToaddr[forsym])

